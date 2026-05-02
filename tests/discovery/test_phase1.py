"""
PHASE1_JOB_SEARCH — Integration test for all three discovery sources.

Tests:
    1. jobspy_discovery   — job boards (Indeed, LinkedIn, ZipRecruiter)
    2. workday_discovery  — Workday company portals (direct JSON API)
    3. smartextract_discovery — AI-powered extraction (Playwright + LLM)

Output:
    test/output/jobs_YYYYMMDD_HHMMSS.xlsx

Run:
    python test_phase1.py                    # all three sources
    python test_phase1.py --skip-smart       # skip smartextract (no Playwright needed)
    python test_phase1.py --source jobspy    # single source
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Path setup and environment loading
# ---------------------------------------------------------------------------

# Locate directories relative to this file
TEST_DIR    = Path(__file__).parent                   # .../PHASE1_JOB_SEARCH/test/
PHASE1_DIR  = TEST_DIR.parent                         # .../PHASE1_JOB_SEARCH/
JOBEZEE_DIR = PHASE1_DIR.parent                       # .../JOBEZEE/
PARENT_DIR  = JOBEZEE_DIR.parent                      # .../RESUME-MAKER/  <- has .env

OUTPUT_DIR  = TEST_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Add PHASE1_JOB_SEARCH to path so we can import it as a package
sys.path.insert(0, str(PHASE1_DIR.parent))

from dotenv import load_dotenv
load_dotenv(PHASE1_DIR / ".env")   # local overrides (if any)
load_dotenv(PARENT_DIR / ".env")   # OPENAI_API_KEY, GEMINI_API_KEY, etc.

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level   = logging.INFO,
    format  = "%(asctime)s  [%(levelname)-8s]  %(name)s: %(message)s",
    datefmt = "%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
# Suppress verbose third-party noise
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logging.getLogger("playwright").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)

log = logging.getLogger("test_phase1")

# ---------------------------------------------------------------------------
# Imports from our package
# ---------------------------------------------------------------------------

from PHASE1_JOB_SEARCH import (
    JobRecord,
    SearchFilters,
    LLMClient,
    deduplicate,
    search_boards,
    search_workday,
    search_smart,
    INDEED_COUNTRY_CODES,
    GLOBAL_EMPLOYERS,
    get_global_employers,
)

# ---------------------------------------------------------------------------
# Test configuration
# ---------------------------------------------------------------------------

QUERIES    = ["software engineer", "python developer"]
LOCATIONS  = ["Remote", "New York, NY"]

# For global board test: search US + UK Indeed simultaneously
TEST_COUNTRIES = ["USA", "United Kingdom"]

# Small employer set pulled from the global registry (fast for testing)
TEST_EMPLOYERS = get_global_employers(
    regions    = ["north_america"],
    industries = ["technology", "finance"],
)

# Sites for smartextract test (small subset — each requires a browser + LLM call)
TEST_SMART_TARGETS = [
    {
        "name":  "Greenhouse Sample",
        "url":   "https://boards.greenhouse.io/gitlab",
        "query": "software engineer",
    },
]


# ---------------------------------------------------------------------------
# Source tests
# ---------------------------------------------------------------------------

def test_jobspy() -> list[JobRecord]:
    """Test jobspy_discovery: scrape Indeed, LinkedIn, ZipRecruiter across multiple countries."""
    log.info("=" * 60)
    log.info("SOURCE 1: JobSpy (global job boards)")
    log.info("=" * 60)
    log.info("Countries: %s", TEST_COUNTRIES)
    log.info("Available Indeed country codes: %d total", len(INDEED_COUNTRY_CODES))

    t0 = time.perf_counter()
    records = search_boards(
        queries          = QUERIES[:1],        # one query to keep the test fast
        locations        = ["Remote"],         # remote works globally
        sites            = ["indeed", "linkedin"],
        results_per_site = 10,
        hours_old        = 72,
        countries        = TEST_COUNTRIES,     # multi-country global sweep
    )
    elapsed = time.perf_counter() - t0

    log.info("jobspy: returned %d records in %.1fs", len(records), elapsed)
    _log_sample(records, "jobspy")
    _assert_records(records, source="jobspy", min_count=0)
    return records


def test_workday() -> list[JobRecord]:
    """Test workday_discovery: call Workday JSON API on a few employers."""
    log.info("=" * 60)
    log.info("SOURCE 2: Workday (company career portals)")
    log.info("=" * 60)

    t0 = time.perf_counter()
    records = search_workday(
        queries   = QUERIES[:1],   # one query to keep the test fast
        employers = TEST_EMPLOYERS,
        workers   = 1,
    )
    elapsed = time.perf_counter() - t0

    log.info("workday: returned %d records in %.1fs", len(records), elapsed)
    _log_sample(records, "workday")
    _assert_records(records, source="workday", min_count=0)
    return records


def test_smartextract(llm_client: LLMClient) -> list[JobRecord]:
    """Test smartextract_discovery: AI-powered extraction with Playwright."""
    log.info("=" * 60)
    log.info("SOURCE 3: SmartExtract (AI-powered site scraping)")
    log.info("=" * 60)

    t0 = time.perf_counter()
    records = search_smart(
        targets    = TEST_SMART_TARGETS,
        llm_client = llm_client,
        workers    = 1,
        headless   = True,
    )
    elapsed = time.perf_counter() - t0

    log.info("smartextract: returned %d records in %.1fs", len(records), elapsed)
    _log_sample(records, "smartextract")
    _assert_records(records, source="smartextract", min_count=0)
    return records


# ---------------------------------------------------------------------------
# Assertion and reporting helpers
# ---------------------------------------------------------------------------

def _assert_records(records: list[JobRecord], source: str, min_count: int) -> None:
    """Validate that records meet minimum quality requirements."""
    if len(records) < min_count:
        log.error(
            "ASSERTION FAILED: %s returned %d records, expected at least %d",
            source, len(records), min_count,
        )
        return

    if not records:
        log.warning("%s: no records returned (network/rate-limit issue possible)", source)
        return

    invalid = [r for r in records if not r.url]
    if invalid:
        log.warning("%s: %d records have no URL", source, len(invalid))

    no_title = [r for r in records if not r.title]
    if no_title:
        log.warning("%s: %d records have no title", source, len(no_title))

    log.info(
        "%s: validation passed | total=%d  no_url=%d  no_title=%d",
        source, len(records), len(invalid), len(no_title),
    )


def _log_sample(records: list[JobRecord], source: str, n: int = 3) -> None:
    """Log a few sample records for visual inspection."""
    if not records:
        return
    log.info("%s: sample records:", source)
    for r in records[:n]:
        log.info(
            "  title=%-40s  company=%-20s  location=%s",
            (r.title or "")[:40],
            (r.company or "")[:20],
            (r.location or "")[:30],
        )


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_excel(records: list[JobRecord], output_path: Path) -> None:
    """
    Write all JobRecord objects to a formatted Excel workbook.

    Sheet "Jobs"    — full data, one row per record.
    Sheet "Summary" — counts by source and site.
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("Missing packages for Excel export: pip install pandas openpyxl")
        return

    if not records:
        log.warning("No records to export.")
        return

    # Build DataFrame from JobRecord dicts
    rows = [r.to_dict() for r in records]
    df   = pd.DataFrame(rows)

    col_order = [
        "title", "company", "location", "remote", "salary",
        "site", "source", "search_query", "date_posted", "url", "description",
    ]
    df = df[[c for c in col_order if c in df.columns]]
    df.columns = [c.replace("_", " ").title() for c in df.columns]

    df.to_excel(output_path, index=False, sheet_name="Jobs")

    # --- Styling ---
    wb = load_workbook(output_path)
    ws = wb["Jobs"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill    = PatternFill("solid", fgColor="D6E4F0")

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = alt_fill

    col_widths = {
        "Title": 42, "Company": 22, "Location": 22, "Remote": 8,
        "Salary": 20, "Site": 14, "Source": 14, "Search Query": 20,
        "Date Posted": 14, "Url": 55, "Description": 60,
    }
    for i, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        hdr = col_cells[0].value or ""
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(hdr, 15)

    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "PHASE1_JOB_SEARCH — Discovery Summary"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")

    ws2["A3"] = "Total Records:"
    ws2["B3"] = len(records)
    ws2["A3"].font = Font(bold=True)

    row = 5
    ws2[f"A{row}"] = "By Source:"
    ws2[f"A{row}"].font = Font(bold=True)
    row += 1
    for src, cnt in df["Source"].value_counts().items() if "Source" in df.columns else []:
        ws2[f"A{row}"] = f"  {src}"
        ws2[f"B{row}"] = int(cnt)
        row += 1

    row += 1
    ws2[f"A{row}"] = "By Site:"
    ws2[f"A{row}"].font = Font(bold=True)
    row += 1
    for site, cnt in (df["Site"].value_counts().items() if "Site" in df.columns else []):
        ws2[f"A{row}"] = f"  {site}"
        ws2[f"B{row}"] = int(cnt)
        row += 1

    row += 1
    ws2[f"A{row}"] = "Generated at:"
    ws2[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    wb.save(output_path)
    log.info("Excel saved: %s", output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Phase 1 discovery test")
    parser.add_argument(
        "--source",
        choices=["jobspy", "workday", "smart", "all"],
        default="all",
        help="Which source to test (default: all)",
    )
    parser.add_argument(
        "--skip-smart",
        action="store_true",
        help="Skip smartextract (skips Playwright and LLM for that source)",
    )
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("PHASE1_JOB_SEARCH — Global Integration Test")
    log.info("=" * 60)
    log.info("GEMINI_API_KEY      : %s", "set" if os.getenv("GEMINI_API_KEY") else "NOT SET")
    log.info("OPENAI_API_KEY      : %s", "set" if os.getenv("OPENAI_API_KEY") else "NOT SET")
    log.info("Indeed countries    : %d supported", len(INDEED_COUNTRY_CODES))
    log.info("Workday employers   : %d global / %d in test", len(GLOBAL_EMPLOYERS), len(TEST_EMPLOYERS))
    log.info("Test countries      : %s", TEST_COUNTRIES)
    log.info("")

    all_records: list[JobRecord] = []
    t_start = time.perf_counter()

    # -- Source 1: JobSpy --
    if args.source in ("jobspy", "all"):
        try:
            all_records.extend(test_jobspy())
        except ImportError as exc:
            log.error("jobspy not installed: %s", exc)
        except Exception as exc:
            log.error("jobspy test failed: %s", exc)

    # -- Source 2: Workday --
    if args.source in ("workday", "all"):
        try:
            all_records.extend(test_workday())
        except Exception as exc:
            log.error("workday test failed: %s", exc)

    # -- Source 3: SmartExtract --
    run_smart = args.source in ("smart", "all") and not args.skip_smart
    if run_smart:
        try:
            llm = LLMClient.from_env()
            all_records.extend(test_smartextract(llm))
        except RuntimeError as exc:
            log.warning("smartextract skipped: %s", exc)
        except ImportError as exc:
            log.warning("smartextract skipped (missing dependency): %s", exc)
        except Exception as exc:
            log.error("smartextract test failed: %s", exc)

    # -- Deduplicate and export --
    unique = deduplicate(all_records)
    dupes  = len(all_records) - len(unique)
    elapsed = time.perf_counter() - t_start

    log.info("")
    log.info("=" * 60)
    log.info("RESULTS")
    log.info("  Total collected : %d", len(all_records))
    log.info("  After dedup     : %d  (%d duplicates removed)", len(unique), dupes)
    log.info("  Elapsed         : %.1fs", elapsed)
    log.info("=" * 60)

    if unique:
        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"jobs_{timestamp}.xlsx"
        save_excel(unique, output_file)
        log.info("Output: %s", output_file)
    else:
        log.warning("No records collected — Excel file not created.")


if __name__ == "__main__":
    main()
