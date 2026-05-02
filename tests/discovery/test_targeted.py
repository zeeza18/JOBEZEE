"""
discovery — Targeted search test
==========================================
Runs two real searches using UserPreferences:

  Search A — USA · AI Engineer · Fresher · Chicago
  Search B — Global Remote · AI Engineer (all levels)

Output:
    tests/discovery/output/targeted_YYYYMMDD_HHMMSS.xlsx
        Sheet "Chicago_AI_Fresher"  — Search A results
        Sheet "Global_Remote_AI"    — Search B results
        Sheet "Summary"             — counts and stats

Run:
    python tests/discovery/test_targeted.py                  # boards + workday
    python tests/discovery/test_targeted.py --boards-only    # skip workday (faster)
    python tests/discovery/test_targeted.py --workday-only   # skip boards
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths + env
# ---------------------------------------------------------------------------
TEST_DIR    = Path(__file__).parent           # .../tests/discovery/
JOBEZEE_DIR = TEST_DIR.parent.parent          # .../JOBEZEE/
PARENT_DIR  = JOBEZEE_DIR.parent              # .../RESUME-MAKER/  (has .env)
OUTPUT_DIR  = TEST_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

sys.path.insert(0, str(JOBEZEE_DIR))

from dotenv import load_dotenv
load_dotenv(JOBEZEE_DIR / ".env")
load_dotenv(PARENT_DIR / ".env")

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level   = logging.INFO,
    format  = "%(asctime)s  [%(levelname)-8s]  %(message)s",
    datefmt = "%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
log = logging.getLogger("test_targeted")

# ---------------------------------------------------------------------------
# Imports
# ---------------------------------------------------------------------------
from discovery import (
    UserPreferences,
    JobRecord,
    deduplicate,
    search_boards,
    search_workday,
)

# ===========================================================================
# SEARCH A — USA · AI Engineer · Fresher · Chicago
# ===========================================================================
SEARCH_A = UserPreferences(
    job_titles = [
        "AI engineer",
        "machine learning engineer",
        "artificial intelligence engineer",
        "ML engineer",
        "deep learning engineer",
    ],
    locations        = ["Chicago, IL", "Remote"],
    countries        = ["USA"],
    remote_ok        = True,         # include remote postings from US employers
    experience_level = "entry",
    industries       = ["technology"],
    hours_old        = 168,          # 7 days — fresh postings for new grads
    results_per_site = 25,
    # drop anything that requires experience or seniority
    exclude_titles   = [
        "senior", "sr.", "lead", "principal", "staff", "director",
        "vp ", "vice president", "manager", "head of", "chief",
        "experienced", "mid-level", "clearance required",
    ],
)

# ===========================================================================
# SEARCH B — Global Remote · AI Engineer (all levels)
# ===========================================================================
SEARCH_B = UserPreferences(
    job_titles = [
        "AI engineer",
        "machine learning engineer",
        "LLM engineer",
        "generative AI engineer",
        "ML engineer",
        "AI researcher",
    ],
    remote_ok        = True,    # global remote — no country/region restriction
    experience_level = "",      # all levels
    industries       = ["technology"],
    hours_old        = 72,      # last 3 days
    results_per_site = 30,
    exclude_titles   = [
        "director", "vp ", "vice president", "chief",
        "clearance required", "ts/sci",
    ],
)


# ===========================================================================
# Runner helpers
# ===========================================================================

def run_boards(label: str, prefs: UserPreferences) -> list[JobRecord]:
    """Run JobSpy job board search for a given UserPreferences."""
    log.info("  [Boards] %s | queries=%s", label, prefs.job_titles)
    t0      = time.perf_counter()
    kwargs  = prefs.to_jobspy_kwargs()
    records = search_boards(
        **kwargs,
        sites    = ["indeed", "linkedin", "zip_recruiter"],
    )
    elapsed = time.perf_counter() - t0

    # Apply title exclusion filter
    before = len(records)
    records = [r for r in records if prefs.to_search_filters().passes_title(r.title)]
    dropped = before - len(records)

    log.info(
        "  [Boards] %s | raw=%d  excluded=%d  kept=%d  (%.1fs)",
        label, before, dropped, len(records), elapsed,
    )
    return records


def run_workday(label: str, prefs: UserPreferences) -> list[JobRecord]:
    """Run Workday API search for a given UserPreferences."""
    employers = prefs.get_workday_employers()
    log.info(
        "  [Workday] %s | queries=%s | employers=%d",
        label, prefs.job_titles[:2], len(employers),
    )
    if not employers:
        log.warning("  [Workday] %s | no employers matched — skipping", label)
        return []

    t0      = time.perf_counter()
    records = search_workday(
        queries               = prefs.job_titles,
        employers             = employers,
        workers               = 4,
        fetch_details         = False,   # skip per-job detail GET — avoids hanging
        max_jobs_per_employer = 30,      # cap per employer per query for speed
    )
    elapsed = time.perf_counter() - t0

    # Apply title exclusion filter
    before = len(records)
    sf     = prefs.to_search_filters()
    records = [r for r in records if sf.passes_title(r.title)]
    dropped = before - len(records)

    log.info(
        "  [Workday] %s | raw=%d  excluded=%d  kept=%d  (%.1fs)",
        label, before, dropped, len(records), elapsed,
    )
    return records


# ===========================================================================
# Excel export
# ===========================================================================

def save_excel(
    results: dict[str, list[JobRecord]],
    output_path: Path,
) -> None:
    """
    Write multiple result sets to one Excel file — one sheet per search.
    Final sheet is a Summary with counts.

    Args:
        results:     {sheet_name: list[JobRecord]}
        output_path: where to write the .xlsx
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("Missing: pip install pandas openpyxl")
        return

    col_order = [
        "title", "company", "location", "remote", "salary",
        "site", "source", "search_query", "date_posted",
        "experience_level", "country", "url", "description",
    ]
    col_widths = {
        "Title": 42, "Company": 22, "Location": 25, "Remote": 8,
        "Salary": 20, "Site": 14, "Source": 14, "Search Query": 22,
        "Date Posted": 14, "Experience Level": 18, "Country": 16,
        "Url": 55, "Description": 60,
    }

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill    = PatternFill("solid", fgColor="D6E4F0")

    # We need a temporary first sheet to write; openpyxl can't start from scratch
    first_sheet = next(iter(results))
    first_df    = _to_df(results[first_sheet], col_order)
    first_df.to_excel(output_path, index=False, sheet_name=first_sheet)

    wb = load_workbook(output_path)

    for sheet_name, records in results.items():
        df = _to_df(records, col_order)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            # write header + data manually
            from openpyxl.utils.dataframe import dataframe_to_rows
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        _style_sheet(ws, header_fill, header_font, alt_fill, col_widths)

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "discovery — Targeted Search Results"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws_sum["A3"] = "Generated at:"
    ws_sum["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row = 5
    for sheet_name, records in results.items():
        ws_sum[f"A{row}"] = sheet_name
        ws_sum[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        ws_sum[f"A{row}"] = "  Total unique jobs:"
        ws_sum[f"B{row}"] = len(records)
        row += 1

        if records:
            from collections import Counter
            by_source = Counter(r.source for r in records)
            by_remote = Counter("Remote" if r.remote else "On-site" for r in records)
            by_site   = Counter(r.site   for r in records)

            ws_sum[f"A{row}"] = "  By source:"
            row += 1
            for src, cnt in by_source.most_common():
                ws_sum[f"A{row}"] = f"    {src}"
                ws_sum[f"B{row}"] = cnt
                row += 1

            ws_sum[f"A{row}"] = "  Remote vs On-site:"
            row += 1
            for k, cnt in by_remote.most_common():
                ws_sum[f"A{row}"] = f"    {k}"
                ws_sum[f"B{row}"] = cnt
                row += 1

            ws_sum[f"A{row}"] = "  Top sites:"
            row += 1
            for site, cnt in by_site.most_common(5):
                ws_sum[f"A{row}"] = f"    {site}"
                ws_sum[f"B{row}"] = cnt
                row += 1

        row += 1   # blank separator between searches

    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 12

    wb.save(output_path)
    log.info("Saved → %s", output_path)


def _to_df(records: list[JobRecord], col_order: list[str]):
    import pandas as pd
    rows = [r.to_dict() for r in records]
    if not rows:
        return pd.DataFrame(columns=[c.replace("_", " ").title() for c in col_order])
    df = pd.DataFrame(rows)
    df = df[[c for c in col_order if c in df.columns]]
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    return df


def _style_sheet(ws, header_fill, header_font, alt_fill, col_widths):
    from openpyxl.styles import Alignment

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = alt_fill

    for i, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        hdr = col_cells[0].value or ""
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(hdr, 15)

    ws.freeze_panes = "A2"


# ===========================================================================
# Main
# ===========================================================================

def main() -> None:
    parser = argparse.ArgumentParser(description="Targeted AI Engineer search test")
    parser.add_argument("--boards-only",  action="store_true", help="Skip Workday")
    parser.add_argument("--workday-only", action="store_true", help="Skip job boards")
    args = parser.parse_args()

    t_start = time.perf_counter()

    log.info("=" * 65)
    log.info("  discovery — Targeted AI Engineer Test")
    log.info("=" * 65)
    log.info("")
    log.info("  Search A : USA · AI Engineer · Fresher · Chicago")
    log.info("  Search B : Global Remote · AI Engineer (all levels)")
    log.info("")

    results: dict[str, list[JobRecord]] = {
        "Chicago_AI_Fresher": [],
        "Global_Remote_AI":   [],
    }

    # ── Search A ────────────────────────────────────────────────────────────
    log.info("━" * 65)
    log.info("SEARCH A — USA · AI Engineer · Fresher · Chicago")
    log.info("━" * 65)

    a_jobs: list[JobRecord] = []

    if not args.workday_only:
        log.info("[1/2] Job boards (Indeed + LinkedIn + ZipRecruiter) …")
        a_jobs.extend(run_boards("Search-A", SEARCH_A))

    if not args.boards_only:
        log.info("[2/2] Workday portals (US tech employers) …")
        a_jobs.extend(run_workday("Search-A", SEARCH_A))

    a_unique = deduplicate(a_jobs)
    results["Chicago_AI_Fresher"] = a_unique

    log.info("")
    log.info("  Search A total : %d unique jobs  (%d dupes removed)",
             len(a_unique), len(a_jobs) - len(a_unique))
    log.info("")
    if a_unique:
        log.info("  Sample results:")
        for r in a_unique[:5]:
            remote_tag = "[Remote]" if r.remote else "[On-site]"
            log.info("    %-45s  %-22s  %s  %s",
                     (r.title or "")[:45],
                     (r.company or "")[:22],
                     remote_tag,
                     (r.location or "")[:25])

    # ── Search B ────────────────────────────────────────────────────────────
    log.info("")
    log.info("━" * 65)
    log.info("SEARCH B — Global Remote · AI Engineer (all levels)")
    log.info("━" * 65)

    b_jobs: list[JobRecord] = []

    if not args.workday_only:
        log.info("[1/2] Job boards (Indeed · LinkedIn · ZipRecruiter — global) …")
        b_jobs.extend(run_boards("Search-B", SEARCH_B))

    if not args.boards_only:
        log.info("[2/2] Workday portals (all 98 global employers) …")
        b_jobs.extend(run_workday("Search-B", SEARCH_B))

    b_unique = deduplicate(b_jobs)
    results["Global_Remote_AI"] = b_unique

    log.info("")
    log.info("  Search B total : %d unique jobs  (%d dupes removed)",
             len(b_unique), len(b_jobs) - len(b_unique))
    log.info("")
    if b_unique:
        log.info("  Sample results (global mix):")
        for r in b_unique[:5]:
            log.info("    %-45s  %-22s  %s",
                     (r.title or "")[:45],
                     (r.company or "")[:22],
                     (r.location or "")[:30])

    # ── Excel export ─────────────────────────────────────────────────────────
    log.info("")
    log.info("━" * 65)
    total     = len(a_unique) + len(b_unique)
    elapsed   = time.perf_counter() - t_start
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file  = OUTPUT_DIR / f"targeted_{timestamp}.xlsx"

    log.info("Saving %d total jobs to Excel …", total)
    save_excel(results, out_file)

    log.info("")
    log.info("=" * 65)
    log.info("  DONE in %.1fs", elapsed)
    log.info("  Search A (Chicago AI Fresher) : %d jobs", len(a_unique))
    log.info("  Search B (Global Remote AI)   : %d jobs", len(b_unique))
    log.info("  Total                         : %d jobs", total)
    log.info("  Output file                   : %s", out_file)
    log.info("=" * 65)


if __name__ == "__main__":
    main()
