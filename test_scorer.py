"""
test_scorer.py — Score every JD+Resume pair in jd_resume_export.xlsx.

Uses the EXACT same scorer as production (score_job_for_user) PLUS shows
each component score and 5 weight configurations side by side.

Output columns:
  A  JD (input)
  B  Resume (input)
  C  JD Title extracted
  D  JD Keywords extracted
  E  Matching skills
  F  Missing skills
  --- component breakdown (W1 default weights) ---
  G  Semantic %   (TF-IDF cosine — language similarity)
  H  Skills %     (fuzzy precision on JD keywords)
  I  Exp %        (years fit)
  J  Title %      (title alignment)
  --- final scores across 5 weight sets ---
  K  W1 FINAL  40s / 25sk / 20e / 15t  ← production default
  L  W2 FINAL  30s / 40sk / 20e / 10t  ← skills-heavy
  M  W3 FINAL  50s / 25sk / 15e / 10t  ← semantic-heavy
  N  W4 FINAL  35s / 35sk / 20e / 10t  ← balanced
  O  W5 FINAL  25s / 45sk / 20e / 10t  ← max skills

Run: python test_scorer.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from backend.services.extraction_service import extract_jd, extract_resume
from backend.services.scorer_service import (
    _semantic_score, _skills_score, _exp_score, _title_score,
    _extract_skills_fallback, _extract_years_from_text, _clean,
)

XLSX_IN  = "jd_resume_export.xlsx"
XLSX_OUT = "jd_resume_scored.xlsx"

# Columns: (label, semantic_w, skills_w, exp_w, title_w)
WEIGHT_SETS = [
    ("W1 sk40/se25/e20/t15", 0.25, 0.40, 0.20, 0.15),  # ← NEW default (skills-first + semantic)
    ("W2 sk50/se20/e20/t10", 0.20, 0.50, 0.20, 0.10),  # skills dominant
    ("W3 sk35/se30/e20/t15", 0.30, 0.35, 0.20, 0.15),  # semantic stronger
    ("W4 sk40/se20/e30/t10", 0.20, 0.40, 0.30, 0.10),  # experience-heavy
    ("W5 sk45/se25/e20/t10", 0.25, 0.45, 0.20, 0.10),  # max skills precision
]

def boost(raw: float) -> int:
    """Apply +10 display boost, cap at 98 — same as frontend."""
    return min(round(raw * 100) + 10, 98)

def pct(f: float) -> str:
    return f"{round(f * 100)}%"

def compute(jd_text, jd_title, jd_skills, jd_years, resume_extraction, resume_raw):
    r_skills = resume_extraction.get("skills") or []
    r_years  = int(resume_extraction.get("years_exp") or 0)
    r_title  = str(resume_extraction.get("title") or "")
    location = resume_extraction.get("location") or ""

    j_skills = list(jd_skills) if jd_skills else _extract_skills_fallback(jd_text)
    j_years  = int(jd_years or 0) or _extract_years_from_text(jd_text)

    sem_text = resume_raw or f"{r_title}. Skills: {', '.join(r_skills)}. {r_years} years. {location}"

    skill_pct, matched, missing = _skills_score(r_skills, j_skills)
    exp_s   = _exp_score(r_years, j_years)
    title_s = _title_score(r_title, jd_title)

    jd_words = len((jd_text or "").split())
    if jd_words < 30:
        sem_s = 0.0
        components = dict(semantic=0.0, skills=skill_pct, exp=exp_s, title=title_s)
        def final(sw, skw, ew, tw):
            return round(skill_pct * 0.50 + exp_s * 0.30 + title_s * 0.20, 4)
    else:
        sem_s = _semantic_score(sem_text, jd_text)
        components = dict(semantic=sem_s, skills=skill_pct, exp=exp_s, title=title_s)
        def final(sw, skw, ew, tw):
            return round(sem_s * sw + skill_pct * skw + exp_s * ew + title_s * tw, 4)

    return components, matched, missing, final

def style_header(cell, color="4472C4"):
    cell.font  = Font(bold=True, color="FFFFFF")
    cell.fill  = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

def check_groq():
    from backend.services.extraction_service import _load_env_key, _GROQ_MODEL
    key = _load_env_key("GROQ_API_KEY")
    if not key:
        print("⚠  GROQ_API_KEY not found — will use regex fallback (lower quality)")
        return False
    print(f"✓  GROQ_API_KEY loaded ({key[:8]}...) | model: {_GROQ_MODEL}")
    return True

def run():
    llm_ok = check_groq()
    print()
    wb = openpyxl.load_workbook(XLSX_IN)
    ws = wb.active

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        (1,  "JD",              "BBBBBB"),
        (2,  "Resume",          "BBBBBB"),
        (3,  "JD Title",        "9FC5E8"),
        (4,  "JD Keywords",     "9FC5E8"),
        (5,  "Matching Skills", "93C47D"),
        (6,  "Missing Skills",  "E06666"),
        (7,  "Semantic %",      "FFD966"),
        (8,  "Skills %",        "FFD966"),
        (9,  "Exp %",           "FFD966"),
        (10, "Title %",         "FFD966"),
    ]
    for col, label, color in headers:
        style_header(ws.cell(row=1, column=col), color)
        ws.cell(row=1, column=col).value = label

    score_colors = ["6AA84F", "3D85C8", "C27BA0", "E69138", "674EA7"]
    for i, (label, *_) in enumerate(WEIGHT_SETS):
        col = 11 + i
        cell = ws.cell(row=1, column=col)
        cell.value = label
        style_header(cell, score_colors[i])

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    for col_letter in "CDEFGHIJ":
        ws.column_dimensions[col_letter].width = 30
    for i in range(len(WEIGHT_SETS)):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(11 + i)].width = 22

    # ── Data ──────────────────────────────────────────────────────────────────
    for row_idx in range(2, ws.max_row + 1):
        jd_text  = str(ws.cell(row=row_idx, column=1).value or "").strip()
        res_text = str(ws.cell(row=row_idx, column=2).value or "").strip()

        if not jd_text or not res_text:
            print(f"  Row {row_idx}: skipped")
            continue

        print(f"Row {row_idx}/{ws.max_row} ...", end=" ", flush=True)

        jd_ext = extract_jd(jd_text)
        jd_skills = jd_ext.skills    if jd_ext else []
        jd_years  = jd_ext.years_min if jd_ext else 0
        jd_title  = jd_ext.title     if jd_ext else ""
        if not jd_ext: print("(JD fallback)", end=" ")

        res_ext = extract_resume(res_text)
        resume_extraction = {
            "title"    : res_ext.title     if res_ext else "",
            "skills"   : res_ext.skills    if res_ext else [],
            "years_exp": res_ext.years_exp if res_ext else 0,
            "education": res_ext.education if res_ext else "none",
            "location" : res_ext.location  if res_ext else "",
        }
        if not res_ext: print("(Resume fallback)", end=" ")

        comps, matched, missing, final_fn = compute(
            jd_text, jd_title, jd_skills, jd_years, resume_extraction, res_text
        )

        ws.cell(row=row_idx, column=3).value = jd_title
        ws.cell(row=row_idx, column=4).value = ", ".join(jd_skills[:50])
        ws.cell(row=row_idx, column=5).value = ", ".join(matched)
        ws.cell(row=row_idx, column=6).value = ", ".join(missing)
        ws.cell(row=row_idx, column=7).value = pct(comps["semantic"])
        ws.cell(row=row_idx, column=8).value = pct(comps["skills"])
        ws.cell(row=row_idx, column=9).value = pct(comps["exp"])
        ws.cell(row=row_idx, column=10).value = pct(comps["title"])

        score_line = []
        for i, (label, sw, skw, ew, tw) in enumerate(WEIGHT_SETS):
            raw  = final_fn(sw, skw, ew, tw)
            disp = boost(raw)
            ws.cell(row=row_idx, column=11 + i).value = f"{disp}%"
            score_line.append(f"{label.split()[0]}={disp}%")

        comp_line = (f"sem={pct(comps['semantic'])} sk={pct(comps['skills'])} "
                     f"exp={pct(comps['exp'])} ttl={pct(comps['title'])}")
        print(f"  [{comp_line}]  " + "  ".join(score_line))

    wb.save(XLSX_OUT)
    print(f"\nSaved → {XLSX_OUT}")
    print("Col G-J = component breakdown | Col K-O = final scores per weight set")
    print("Pick the column (K/L/M/N/O) that feels most accurate → tell me the letter.")

if __name__ == "__main__":
    run()
